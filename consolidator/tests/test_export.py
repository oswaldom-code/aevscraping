"""Verifica el export a .xlsx y la partición por tamaño. Requiere openpyxl
(correr con el venv). Hermético: escribe en un directorio temporal."""

import tempfile
import unittest
from pathlib import Path

import openpyxl

from consolidator.lib.export_xlsx import export_delta


def make(n, pad=""):
    return [{"id": str(i), "nombre": f"Persona {i}", "fuente": "f",
             "edad": i % 90, "es_menor": (i % 2 == 0),
             "observaciones": pad} for i in range(n)]


def total_rows(paths):
    total = 0
    for p in paths:
        wb = openpyxl.load_workbook(p, read_only=True)
        total += sum(1 for _ in wb.active.iter_rows()) - 1  # menos el header
        wb.close()
    return total


class TestExport(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.out = Path(self.tmp.name)

    def tearDown(self):
        self.tmp.cleanup()

    def test_empty_generates_no_files(self):
        self.assertEqual(export_delta([], self.out), [])

    def test_single_file_rows_intact(self):
        paths = export_delta(make(10), self.out, max_mb=5)
        self.assertEqual(len(paths), 1)
        self.assertEqual(total_rows(paths), 10)
        # encabezado correcto
        wb = openpyxl.load_workbook(paths[0], read_only=True)
        header = [c.value for c in next(wb.active.iter_rows())]
        wb.close()
        self.assertEqual(header[0], "ID")
        self.assertEqual(header[-1], "Fuente")

    def test_partition_preserves_rows(self):
        # filas con texto largo + tope pequeño -> fuerza varias partes
        paths = export_delta(make(800, pad="x" * 400), self.out, max_mb=0.03)
        self.assertGreaterEqual(len(paths), 2)
        self.assertEqual(total_rows(paths), 800)


if __name__ == "__main__":
    unittest.main()
